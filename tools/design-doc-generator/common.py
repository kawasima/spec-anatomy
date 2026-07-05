"""Nablarch開発標準の空テンプレートに出張申請システムの内容を流し込むための共通ヘルパ。

テンプレート（examples/business-trip/traditional-design/_templates/）は Nablarch開発標準
（CC BY-SA 4.0）の空書式そのもの。ここでは書式に手を触れず、セル値だけを書き込む。
"""
from __future__ import annotations

import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import (
    column_index_from_string,
    coordinate_to_tuple,
    get_column_letter,
)

# リポジトリ内の相対パス基準
_HERE = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.abspath(os.path.join(_HERE, "..", ".."))
TEMPLATES_DIR = os.path.join(
    REPO_ROOT, "examples", "business-trip", "traditional-design", "_templates"
)
OUTPUT_DIR = os.path.join(REPO_ROOT, "examples", "business-trip", "traditional-design")

# この設計書一式が対象とする（架空の）プロジェクト・システムの識別情報。
# Nablarch書式のヘッダ（PJ名 / システム名 / サブシステム名 / 作成・変更）に流し込む。
PROJECT = {
    "pj": "出張旅費システム構築",
    "system": "出張旅費システム",
    "subsystem": "出張申請",
    "subsystem_id": "BT",
    "author": "アナトミアSI",
    "created": "2024-06-03",
    "changed": "2024-09-17",
}


def template_path(filename: str) -> str:
    return os.path.join(TEMPLATES_DIR, filename)


def output_path(filename: str) -> str:
    return os.path.join(OUTPUT_DIR, filename)


def ensure_template(filename: str) -> str:
    """テンプレートが _templates/ に無ければ Nablarch開発標準（CC BY-SA 4.0）から取得する。

    _templates/ は .gitignore 済み（他者のCC BY-SA空フォーマットをリポジトリに同梱しない）。
    取得元は templates_manifest.py の対応表。ネットワークが無い初回はここで失敗する。
    """
    path = template_path(filename)
    if os.path.exists(path):
        return path
    import urllib.parse
    import urllib.request
    from templates_manifest import DOC_ROOT, GH_BASE, SOURCES

    src = SOURCES.get(filename)
    if src is None:
        raise FileNotFoundError(
            f"テンプレート '{filename}' が _templates/ に無く、取得元も templates_manifest.py に未登録です。"
        )
    # ホストはASCII、パス（非ASCII含む）はまとめてURLエンコード（"/" は区切りとして残す）
    url = GH_BASE + urllib.parse.quote(DOC_ROOT + src)
    os.makedirs(TEMPLATES_DIR, exist_ok=True)
    print(f"  取得: {filename} <- Nablarch開発標準")
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 (spec-anatomy build)"})
    with urllib.request.urlopen(req) as resp, open(path, "wb") as f:
        f.write(resp.read())
    return path


def open_template(filename: str):
    """空テンプレートを（書式を保ったまま）開く。無ければ取得する。"""
    return load_workbook(ensure_template(filename))


def _anchor_coord(ws: Worksheet, coord: str) -> str:
    """coord が結合セルの内側なら、その結合範囲の左上（書き込み可能なセル）を返す。"""
    row, col = coordinate_to_tuple(coord)
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return f"{rng.coord.split(':')[0]}"
    return coord


def put(ws: Worksheet, coord: str, value):
    """結合セルを考慮して安全に値を書き込む。"""
    if value is None:
        return
    ws[_anchor_coord(ws, coord)] = value


def fill_history(wb, product_name: str | None = None):
    """変更履歴シートにプロジェクト識別と変更履歴を記入する（全ブック共通の正準処理）。

    Nablarch書式では、各コンテンツシートの1〜3行目ヘッダ（PJ名 / システム名 /
    サブシステム名 / 成果物名 / 作成・変更）はすべて `=INDIRECT("変更履歴!…")` の数式で、
    値の実体は変更履歴シートにある。従って値の書き込み先は変更履歴シートに一本化する。
    作成・変更の会社／日付は、変更履歴の明細（8行目＝初版、9行目＝変更）から数式で導出される。
    """
    ws = wb["変更履歴"]
    put(ws, "E1", PROJECT["pj"])
    put(ws, "E2", PROJECT["system"])
    put(ws, "E3", PROJECT["subsystem"])
    if product_name is not None:
        put(ws, "S1", product_name)
    # 変更履歴 明細（列: A=No / B=版数 / D=変更日 / G=区分 / J=変更箇所 / Q=変更内容 / AF=担当者）
    rows = [
        (8, 1, "1.00", PROJECT["created"], "新規", "-", "初版作成", PROJECT["author"]),
        (9, 2, "1.01", PROJECT["changed"], "修正", "-", "レビュー指摘反映", PROJECT["author"]),
    ]
    for r, no, ver, date, kind, place, desc, author in rows:
        put(ws, f"A{r}", no)
        put(ws, f"B{r}", ver)
        put(ws, f"D{r}", date)
        put(ws, f"G{r}", kind)
        put(ws, f"J{r}", place)
        put(ws, f"Q{r}", desc)
        put(ws, f"AF{r}", author)


def fill_header(ws: Worksheet, product_name: str | None = None):
    """後方互換のエイリアス。どのシートを渡されても変更履歴シートへ集約する。

    コンテンツシートの1〜3行目ヘッダは変更履歴を参照するINDIRECT数式なので、直接書くと
    数式を潰してしまう。ここでは渡されたシートの親ブックを辿り fill_history に委譲する。
    """
    fill_history(ws.parent, product_name)


def fill_table(ws: Worksheet, start_row: int, colmap: dict[str, str], records: list[dict]):
    """表形式シートに行を流し込む。

    colmap: {列文字: レコードのキー} 例 {"C": "no", "D": "code_id", "P": "value"}
    records: 各行の dict のリスト。キーが無い列は空のまま。
    """
    for i, rec in enumerate(records):
        r = start_row + i
        for col, key in colmap.items():
            if key in rec and rec[key] is not None:
                put(ws, f"{col}{r}", rec[key])


def set_print_area(ws, right_col=35, min_row_floor=1):
    """シートの print_area を A1:{right}{最終内容行} に設定する。

    複製で print_area が消えたシートや、内容が既定印刷範囲をはみ出したシートを救済する。
    right_col はそのテンプレの印刷右端の列番号（画面/バッチ機能設計書なら 35=AI）。
    """
    last = 1
    for row in ws.iter_rows():
        for c in row:
            if c.value not in (None, ""):
                if c.row > last:
                    last = c.row
    ws.print_area = f"A1:{get_column_letter(right_col)}{max(last, min_row_floor)}"


def write_detail_region(ws, rows, start_row, limit=None, right_col=35):
    """自由記述領域に rowdict を上から書く。各セルを次のキー列(無ければ right_col)まで
    横結合し、wrap_text と行高を設定して、狭いセルでの縦折り返しを防ぐ。

    rows: 各行の {列文字: 値} の dict のリスト。空 dict は空行として1行送る。
    start_row: 記述開始行。limit: これを超えたら後続領域に達したとみなし assert で止める。
    right_col: 各行の最右キーを結合する右端の列番号（方眼紙の内容右端に合わせる）。
    """
    # 列幅からおおよその1行あたり全角文字数を出す（方眼紙の既定列幅）
    default_w = (ws.sheet_format.defaultColWidth or 2.5)
    r = start_row
    for rowdict in rows:
        if limit is not None:
            assert r <= limit, "自由記述領域が後続領域に達した"
        keys = sorted(rowdict.items(), key=lambda kv: column_index_from_string(kv[0]))
        max_lines = 1
        for i, (col, val) in enumerate(keys):
            c0 = column_index_from_string(col)
            c1 = (column_index_from_string(keys[i + 1][0]) - 1) if i + 1 < len(keys) else right_col
            if c1 < c0:
                c1 = c0
            put(ws, f"{col}{r}", val)             # 値は左上セルへ
            # この行の該当範囲に残る単一行結合を解除してから結合
            for rng in list(ws.merged_cells.ranges):
                if rng.min_row == r and rng.max_row == r and not (rng.max_col < c0 or rng.min_col > c1):
                    ws.unmerge_cells(str(rng))
            if c1 > c0:
                ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c1)
            anchor = ws.cell(row=r, column=c0)
            anchor.alignment = Alignment(wrap_text=True, horizontal="left", vertical="top")
            if isinstance(val, str) and val:
                per_line = max(4, int((c1 - c0 + 1) * default_w / 2.1))  # 全角概算
                lines = -(-len(val) // per_line)
                max_lines = max(max_lines, lines)
        cur = ws.row_dimensions[r].height or 15
        ws.row_dimensions[r].height = max(cur, max_lines * 14 + 2)
        r += 1
    return r


def save(wb, filename: str) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out = output_path(filename)
    wb.save(out)
    return out
