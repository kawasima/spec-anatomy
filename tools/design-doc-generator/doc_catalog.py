"""設計書一覧（カタログ） — 出張申請システムの設計書一式の索引。

Nablarch開発標準の「設計書一覧」書式（2.2. 設計書一覧シート）をそのまま流用し、
本システム（出張旅費システム／出張申請サブシステム）で実際に作成した設計書に対して
「提供」列に○を、作成していない設計書に－を記入して索引化する。カタログの設計書名
（テンプレ既定の項目名）はそのまま活かし、取引ID・電文ID・コンポーネントID等の
具体的な作成成果物は「説明」列に追記する。

「提供」列はテンプレ既定では Nablarch開発標準が書式を提供するか否か（あり／なし）を
表していたが、本索引では「本システムで当該設計書を作成したか否か」に読み替える。
"""
from __future__ import annotations

from openpyxl.utils import column_index_from_string

from common import open_template, fill_header, put, save

TEMPLATE = "設計書一覧.xlsx"
OUTPUT = "設計書一覧.xlsx"
SHEET = "2.2. 設計書一覧"

# カタログ表の列（結合なし・単一セル）
COL_NAME = "D"   # 設計書名
COL_DESC = "X"   # 説明
COL_UNIT = "AI"  # 作成単位
COL_PROV = "AK"  # 提供（本索引では「本システムでの作成有無」に読み替え）

# 本システムで作成した設計書（カタログ行 → 具体的な作成成果物の追記文）。
# 追記文が None の行は○のみ（追記不要）。ここに無い行は未作成（－）。
PRODUCED: dict[int, str | None] = {
    28: "本システム: BT 出張申請",  # システム処理フロー
    29: "本システム: BT 出張申請",
    31: "本システム: BTB01 経理連携ファイル作成",
    33: "本システム: BTS01 出張申請登録／BTS02 事前承認／"
        "BTS03 出張実績登録／BTS04 最終承認／BTS05 出張申請一覧",
    34: "本システム: BT 出張申請",
    36: "本システム: BT 出張申請",
    37: "本システム: CC0001 事前承認要否判定／CC0002 精算額計算／"
        "CC0003 経理連携データ作成／CC0004 申請ID採番",
    40: "本システム: BT 出張申請",
    42: "本システム: IBT01 経理システム連携ファイル",
    47: "本システム: BT 出張申請",
    48: "本システム: BT 出張申請",  # 画面遷移図
    50: "本システム: BT 出張申請",
    51: "本システム: RBT01 出張精算書",  # 帳票設計書
    53: "本システム: BT 出張申請",
    54: "本システム: BT 出張申請（MBT01 事前承認依頼／"
        "MBT02 事前承認却下通知／MBT03 最終承認依頼）",
    56: "本システム: BT 出張申請",
    58: None,  # コード設計書（システム）
    61: None,  # テーブル一覧（システム）
    62: "本システム: BT 出張申請"
        "（BT_APPLICATION／BT_PLANNED_COST／BT_ACTUAL_COST／BT_EMPLOYEE）",
    63: None,  # ドメイン定義書（システム）
    64: "本システム: BT 出張申請",
    66: "本システム: NBT0601 経理連携（BTB01 経理連携ファイル作成バッチ）",  # ネット・ジョブフロー
}

# カタログのキャプション（本索引としての読み替えを明記）
CAPTION_CELL = "C6"
CAPTION = (
    "出張旅費システム（出張申請サブシステム）で作成した設計書の一覧を以下に示す。"
    "「提供」列の○は本システムで作成した設計書、－は本システムでは作成対象外の設計書を表す。"
)


def _data_rows(ws) -> list[int]:
    """設計書名（D列）が入っているカタログ明細行の行番号を返す。"""
    rows = []
    for r in range(13, ws.max_row + 1):
        if ws[f"{COL_NAME}{r}"].value not in (None, ""):
            rows.append(r)
    return rows


def build() -> str:
    wb = open_template(TEMPLATE)
    ws = wb[SHEET]

    # 共通ヘッダ（変更履歴シート経由）。本ブックの成果物名は「設計書一覧」。
    fill_header(ws, product_name="設計書一覧")
    # テンプレには Nablarch開発標準自身の変更履歴が3行入っている。本システムの索引としては
    # fill_header が書く2行（初版・レビュー反映）のみとし、残る旧履歴行（10行目）を消す。
    hist = wb["変更履歴"]
    for col in ("A", "B", "D", "G", "J", "Q", "AF"):
        put(hist, f"{col}10", None)
        hist[f"{col}10"] = None

    # カタログのキャプションを本索引向けに差し替え
    put(ws, CAPTION_CELL, CAPTION)

    # 明細行ごとに「提供」列へ○／－を記入し、作成した設計書は説明へ具体的な成果物を追記
    for r in _data_rows(ws):
        if r in PRODUCED:
            put(ws, f"{COL_PROV}{r}", "○")
            note = PRODUCED[r]
            if note:
                anchor = f"{COL_DESC}{r}"
                base = ws[anchor].value
                base = "" if base in (None, "") else str(base)
                put(ws, anchor, f"{base}【{note}】")
        else:
            put(ws, f"{COL_PROV}{r}", "－")

    return save(wb, OUTPUT)


if __name__ == "__main__":
    print("wrote", build())
